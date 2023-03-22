import unicodedata
from django.shortcuts import render
from .models import Compte, Contributeur, Tweet, Corpus, Type_Annotation, Categorie_Sentiment, Annotation, Est_Contradictoire
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from .forms import UploadFileForm
from openpyxl import load_workbook
import xlwt
import snscrape.modules.twitter as sntwitter
from django.utils.encoding import *
from django.contrib.auth.hashers import make_password, check_password

#Login & Sign Up Page
def login(request):
    return render(request, "app_annotation/login.html")
#Username already exists
def login_user_exist(request):
    return render(request, "app_annotation/login.html", {"user_exist": True})
#Registration Successful
def login_after_registration(request):
    return render(request, "app_annotation/login.html", {"signup_success": True})
def login_user_not_found(request):
    return render(request, "app_annotation/login.html", {"user_not_found": True})
def login_wrong_password(request):
    return render(request, "app_annotation/login.html", {"wrong_password": True})
def account_blocked(request):
    return render(request, "app_annotation/login.html", {"account_blocked": True})
def authenticate(request):
    login = request.POST['logusername']
    password = request.POST['logpassword']
    try:
        compte = Compte.objects.get(pk=login)
    except Compte.DoesNotExist:
        return HttpResponseRedirect(reverse("login_user_not_found"))
    if check_password(password, compte.password) and compte.est_admin:
        request.session["login_admin"] = compte.login
        request.session["password_admin"] = compte.password
        return HttpResponseRedirect(reverse("admin_dashboard"))
    if check_password(password, compte.password) and not compte.est_admin:
        compte_cont = Contributeur.objects.get(pk=compte)
        if not compte_cont.bloque:
          request.session["login_contri"] = compte.login
          request.session["password_contri"] = compte.password
          return HttpResponseRedirect(reverse("contributeur_dashboard"))
        else:
          return HttpResponseRedirect(reverse("account_blocked"))

    else:
        return HttpResponseRedirect(reverse("login_wrong_password"))


def dashboard_admin(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        tous_corpus = Corpus.objects.all()
        tous_cont = Contributeur.objects.all()
        corpus_annot = 0
        for corpus in tous_corpus:
            nb_contr = 0
            for cont in Contributeur.objects.all():
                contribution = Annotation.objects.filter(contributeur = cont, corpus = corpus)
                if contribution:
                    nb_contr += 1
            if (nb_contr == len(tous_cont)):
                corpus_annot +=1

        corpus_non_annot = len(tous_corpus) - corpus_annot   
   
        return render(request, "app_annotation/dashboard_admin.html", {"tous_corpus" : tous_corpus, "compte" : compte_session, "corpus_annot": corpus_annot, "tous_cont" : len(tous_cont), "corpus_non_annot": corpus_non_annot})
    else:
        return HttpResponseRedirect(reverse("login"))  

def dashboard_contributeur(request):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:

        tous_corpus = Corpus.objects.all()
        tous_annot = Annotation.objects.filter(contributeur=compte_session)
        list = []
        for corpus in tous_corpus:
            nb_vote = 0
            annotations = Annotation.objects.filter(corpus=corpus, contributeur = compte_session)
            for annot in annotations:
                    nb_vote += 1
            list.append(nb_vote)
       
        return render(request, "app_annotation/dashboard_contributeur.html", {"compte" : compte_session, "tous_corpus": tous_corpus,"tous_annot": tous_annot, "list" : list})
    else:
        return HttpResponseRedirect(reverse("login"))

def register(request):
    nom_post = request.POST['lognom']
    prenom_post = request.POST['logprenom']
    username_post = request.POST['logusername']
    password_post = request.POST['logpassword']
    encryptedpassword = make_password(password_post)
    comptes = Compte.objects.all()
    for compte in comptes:
        if(compte.login== username_post):
            return HttpResponseRedirect(reverse("login_user_exist"))
    compte = Compte(login=username_post, password=encryptedpassword) 
    compte.save() 
    compte_contr = Contributeur(login=compte, nom=nom_post, prenom = prenom_post)  
    compte_contr.save()
    return HttpResponseRedirect(reverse("login_after_registration"))

def logout_contri(request):
    try:
        del request.session["login_contri"]
        del request.session["password_contri"]
    except KeyError:
        return HttpResponseRedirect(reverse("login"))
    return HttpResponseRedirect(reverse("login"))

def logout_admin(request):
    try:
        del request.session["login_admin"]
        del request.session["password_admin"]
    except KeyError:
        return HttpResponseRedirect(reverse("login"))
    return HttpResponseRedirect(reverse("login"))

def upload_corpus(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        corpus_titre = request.POST["corpus_titre"]
        corpus_file = request.FILES['file']
        masque = request.POST["masque"]
        type_annot_post = request.POST["type_annot"]
        type_annot = Type_Annotation.objects.get(pk=(int)(type_annot_post))
        try: 
          workbook = load_workbook(filename=corpus_file.file)
          Page = workbook['Sheet1']
          column = Page["A"]
        except:
          form_redir = UploadFileForm()
          types_annot_redir = Type_Annotation.objects.all()
          return HttpResponseRedirect(reverse("upload_corpus_error"))
          return render(request, "app_annotation/gerer_corpus/upload.html", {"form": form_redir,  "types_annot": types_annot_redir, "error": True})
        corpus = Corpus(titre_corpus=corpus_titre, type_annotation=type_annot, masque=(int)(masque))
        corpus.save()
        for cell in column:
            if cell.value!=None:
              cell_txt =cell.value
              tweet = Tweet(text=cell_txt, corpus=corpus)
              tweet.save()
            else:
                continue
        form_redir = UploadFileForm()
        types_annot_redir = Type_Annotation.objects.all()
        return HttpResponseRedirect(reverse("upload_corpus_done"))
      else:
        form = UploadFileForm()
        types_annot = Type_Annotation.objects.all()
        return render(request, "app_annotation/gerer_corpus/upload.html", {"form": form,  "types_annot": types_annot})
    else:
        return HttpResponseRedirect(reverse("login"))  



def creer_corpus(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        corpus_titre = request.POST["corpus_titre"]
        masque = request.POST["masque"]
        type_annot_post = request.POST["type_annot"]
        type_annot = Type_Annotation.objects.get(pk=(int)(type_annot_post))
        tweets_post = request.POST.getlist("tweet")
        corpus = Corpus(titre_corpus=corpus_titre, type_annotation=type_annot, masque=(int)(masque))
        corpus.save()

        for tweet in tweets_post:
            tweet_db = Tweet(text=tweet, corpus=corpus)
            tweet_db.save()
        types_annot_redir = Type_Annotation.objects.all()
        cree = True
        return HttpResponseRedirect(reverse("creer_corpus_done"))
      else:
        types_annot = Type_Annotation.objects.all()
        return render(request, "app_annotation/gerer_corpus/creer_corpus.html", {"types_annot": types_annot})
    else:
        return HttpResponseRedirect(reverse("login"))  
    
def modifier_corpus(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
           
           tous_corpus = Corpus.objects.all()

           return render(request, "app_annotation/gerer_corpus/modifier_corpus.html", {"tous_corpus": tous_corpus})
        
    else:
        return HttpResponseRedirect(reverse("login"))  
    
def modifier_corpus_id(request, corpus_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
           if request.method == "POST":
              try:  
                corpus = Corpus.objects.get(pk=corpus_id)
                corpus_titre = request.POST["corpus_titre"]
                type_annot_post = request.POST["type_annot"]
                type_annot = Type_Annotation.objects.get(pk=type_annot_post)
                type_annot_old = corpus.type_annotation
                tweets_post = request.POST.getlist("tweet")
                corpus_old_tweets = corpus.tweet_set.all()

                if (int)(type_annot_post) != type_annot_old.id:
                        for tweet in corpus_old_tweets:
                            tweet.delete()
                        for tweet in tweets_post:
                            new_tw = Tweet(text=tweet, corpus=corpus)
                            new_tw.save()                            
                else:

                  found = False
                  for tweet in  corpus_old_tweets:
                    found = False
                    for tw_post in tweets_post:
                        if tw_post == tweet.text:
                            found = True
                    
                    if not found:
                        tweet.delete()
                    else:
                        continue
          

                  for tweet in tweets_post:
                  
                    if not Tweet.objects.filter(text=tweet, corpus = corpus):
                        new_tw = Tweet(text=tweet, corpus=corpus)
                        new_tw.save()

                corpus.type_annotation = type_annot
                corpus.titre_corpus = corpus_titre
                corpus.save()

   
                return HttpResponseRedirect(reverse("modifier_corpus_id_done", args=(corpus_id,)))
              except:
                return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')

           else:
              try: 
               corpus = Corpus.objects.get(pk=corpus_id)
               types_annot = Type_Annotation.objects.all()
               return render(request, "app_annotation/gerer_corpus/modifier_corpus_id.html", {"corpus": corpus, "types_annot": types_annot})
              except:
                return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
        
    else:
        return HttpResponseRedirect(reverse("login")) 


def masquer_corpus(request, corpus_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
           
           corpus = Corpus.objects.get(pk=corpus_id)
           if corpus.masque:
               corpus.masque = False
               corpus.save()
           else:
               corpus.masque = True
               corpus.save()
           return HttpResponseRedirect(reverse("modifier_corpus_id", args=(corpus_id,)))
        
    else:
        return HttpResponseRedirect(reverse("login"))
    
def supprimer_corpus_id(request, corpus_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        corpus = Corpus.objects.get(pk=corpus_id)
        corpus.delete()
        return HttpResponseRedirect(reverse("corpus"))  

    else:
        return HttpResponseRedirect(reverse("login"))  
    
def ajouter_type_annotation(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        type_annotation_post = request.POST["type_annotation"]
        categories = request.POST.getlist("categorie")
        type_annotation = Type_Annotation(nom_type=type_annotation_post)
        type_annotation.save()
        for cat in categories:
            cat_db = Categorie_Sentiment(nom_categorie=cat, type_annotation=type_annotation)
            cat_db.save()
        return HttpResponseRedirect(reverse("ajouter_type_annotation_done")) 
      else:
        return render(request, "app_annotation/gerer_type_annot/ajouter_type_annotation.html")
    else:
        return HttpResponseRedirect(reverse("login")) 



def modifier_type_annotation(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:

            types_annot = Type_Annotation.objects.all()
            return render(request, "app_annotation/gerer_type_annot/modifier_type_annotation.html", {"types_annot": types_annot})
    
        
    else:
        return HttpResponseRedirect(reverse("login"))  


def modifier_type_annotation_id(request, type_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        if request.method == "POST":
          try:
            type_annotation_post = request.POST["type_annotation"]
            categories = request.POST.getlist("categorie")
            type_annot = Type_Annotation.objects.get(pk=type_id)
            
            for cat in type_annot.categorie_sentiment_set.all():
                cat.delete()
            for cat in categories:
                cat_db = Categorie_Sentiment(nom_categorie=cat, type_annotation=type_annot)
                cat_db.save()
            

            type_annot.nom_type = type_annotation_post
            type_annot.save()

            type_annot_red = Type_Annotation.objects.get(pk=type_id)
            categories_red = type_annot.categorie_sentiment_set.all()
            return HttpResponseRedirect(reverse("modifier_type_annotation_id_done", args=(type_id, )))
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')

        else:
          try: 
            type_annot = Type_Annotation.objects.get(pk=type_id)
            categories = type_annot.categorie_sentiment_set.all()
            return render(request, "app_annotation/gerer_type_annot/modifier_type_annotation_id.html", {"type_annot": type_annot, "categories": categories})
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
        
    else:
        return HttpResponseRedirect(reverse("login"))   
    
def supprimer_type_annotation_id(request, type_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        type_annot = Type_Annotation.objects.get(pk=type_id)
        type_annot.delete()
        return HttpResponseRedirect(reverse("type_annotation"))
        
    else:
        return HttpResponseRedirect(reverse("login")) 

def ajouter_etiquette_contradictoire(request, type_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        if request.method == 'POST':
          try:
            type_annot = Type_Annotation.objects.get(pk=type_id)
            old_etiq_contr = type_annot.est_contradictoire_set.all()
            etiq_contr = request.POST.getlist("etiq_contr")
            for old_etiq_conte in old_etiq_contr:
                old_etiq_contr.delete()
            for i in range(0, len(etiq_contr), 2):
                print(etiq_contr[i])
                cat1 = Categorie_Sentiment.objects.get(pk=(int)(etiq_contr[i]))
                cat2 = Categorie_Sentiment.objects.get(pk=(int)(etiq_contr[i+1]))
                contradiction = Est_Contradictoire(cat1=cat1, cat2=cat2, type_annotation= type_annot)
                contradiction.save()
            return HttpResponseRedirect(reverse("ajouter_etiquette_contradictoire_done", args=(type_id, )))
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
        else:
           try: 
            type_annot = Type_Annotation.objects.get(pk=type_id)
            categories = type_annot.categorie_sentiment_set.all()
            etiq_contradictoire = type_annot.est_contradictoire_set.all()
            return render(request, "app_annotation/ajouter_etiquette_contra.html", {"categories": categories, "type_id":type_id, "etiq_contradictoire":  etiq_contradictoire})
           except:
               return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
        
    else:
        return HttpResponseRedirect(reverse("login")) 
    

def ajouter_etiquette_contradictoire_done(request, type_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        if request.method == 'POST':
            type_annot = Type_Annotation.objects.get(pk=type_id)
            old_etiq_contr = type_annot.est_contradictoire_set.all()
            etiq_contr = request.POST.getlist("etiq_contr")
            for old_etiq_conte in old_etiq_contr:
                old_etiq_contr.delete()
            for i in range(0, len(etiq_contr), 2):
                print(etiq_contr[i])
                cat1 = Categorie_Sentiment.objects.get(pk=(int)(etiq_contr[i]))
                cat2 = Categorie_Sentiment.objects.get(pk=(int)(etiq_contr[i+1]))
                contradiction = Est_Contradictoire(cat1=cat1, cat2=cat2, type_annotation= type_annot)
                contradiction.save()
            return HttpResponseRedirect(reverse("ajouter_etiquette_contradictoire_done", args=(type_id, )))
        else:
            type_annot = Type_Annotation.objects.get(pk=type_id)
            categories = type_annot.categorie_sentiment_set.all()
            etiq_contradictoire = type_annot.est_contradictoire_set.all()
            return render(request, "app_annotation/ajouter_etiquette_contra.html", {"categories": categories, "type_id":type_id, "etiq_contradictoire":  etiq_contradictoire, "cree": True})
        
    else:
        return HttpResponseRedirect(reverse("login")) 

def ajouter_etiquette(request):
        
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
            if request.method == "POST":
                nom = request.POST["etiquette"]
                type_post = request.POST["type_annot"]
                type_annot = Type_Annotation.objects.get(pk=type_post)
                etiquette = Categorie_Sentiment(nom_categorie=nom, type_annotation = type_annot)
                etiquette.save()
                types_annot_red = Type_Annotation.objects.all()
                return HttpResponseRedirect(reverse("ajouter_etiquette_done"))
            else:
        
               types_annot = Type_Annotation.objects.all()
               return render(request, "app_annotation/gerer_type_annot/ajouter_etiquette.html", {"types_annot": types_annot})
        
    else:
        return HttpResponseRedirect(reverse("login")) 

    #Contributeur


def afficher_corpus_contr(request):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
        if request.method == "POST":
            pass
        else:
            tous_corpus = Corpus.objects.all()
            tous_annot = Annotation.objects.filter(contributeur=compte_session)
             
            return render(request, "app_annotation/consulter_corpus/afficher_corpus_disp.html", {"tous_corpus": tous_corpus,"tous_annot": tous_annot})
    else:
        return HttpResponseRedirect(reverse("login"))
    

def annoter_corpus(request, corpus_id):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
       if request.method == "POST":
          try: 
           corpus = Corpus.objects.get(pk=corpus_id)
           tweets = corpus.tweet_set.all()
           tous_avis = request.POST.getlist("avis") 
           
           tous_avis_old= Annotation.objects.filter(corpus= corpus, contributeur = compte_session)
           for avis in tous_avis_old:
               avis.delete()
           
           for i, avis in enumerate(tous_avis):
              if avis != "":
                   print(avis)
                   avis_split = avis.split("_")
                   cat_id = avis_split[0]
                   tw_id = avis_split[1]
                   tw = Tweet.objects.get(pk=(int)(tw_id))
                   annot = Annotation(avis=cat_id, tweet = tw, contributeur= compte_session, corpus=corpus)
                   annot.save()
           


           corpus_red = Corpus.objects.get(pk=corpus_id)
           tweets_red = corpus_red.tweet_set.all()
           tous_avis_red = Annotation.objects.filter(corpus=corpus_red, contributeur=compte_session)
           type_annot_red = corpus_red.type_annotation
           categories_red = type_annot_red.categorie_sentiment_set.all()
           return HttpResponseRedirect(reverse("annoter_corpus_done", args=(corpus_id, )))
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
       else:
          try: 
           corpus = Corpus.objects.get(pk=corpus_id)
           tweets = corpus.tweet_set.all()
           tous_avis = Annotation.objects.filter(corpus=corpus, contributeur=compte_session)
           type_annot = corpus.type_annotation
           categories = type_annot.categorie_sentiment_set.all()
           return render(request, "app_annotation/consulter_corpus/annoter_corpus.html", {"corpus": corpus, "categories": categories, "tweets": tweets, "tous_avis": tous_avis, "contributeur_login": compte_session.login.login})
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')


    else:
        return HttpResponseRedirect(reverse("login"))
    
def supprimer_annotation(request, corpus_id, contributeur):

    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
            
            contributeur = Contributeur.objects.get(pk=contributeur)
            corpus = Corpus.objects.get(pk=corpus_id)

            tous_annot = Annotation.objects.filter(corpus= corpus, contributeur = contributeur)

            for annot in tous_annot:
                annot.delete()
            
            return HttpResponseRedirect(reverse("annoter_corpus", args=(corpus_id,)))
    else:
        return HttpResponseRedirect(reverse("login"))
    
  #Admin Resultats

def afficher_corpus_res(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        tous_corpus = Corpus.objects.all()

        return render(request, "app_annotation/consulter_resultats/afficher_corpus.html", {"tous_corpus": tous_corpus})
    else:
        return HttpResponseRedirect(reverse("login"))  
    
def corpus_resultat(request, corpus_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
     
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        categories = corpus.type_annotation.categorie_sentiment_set.all()
        list1 = dict()
        list2 = dict()

   
        for tweet in tweets:
            tous_annnot= Annotation.objects.filter(tweet=tweet)
            list2 = dict()
            for cat in categories:
                votes =0
                for avis in tous_annnot:
                    if avis.avis == cat.id:
                        votes += 1
                list2[cat.id] = votes
            list1[tweet.id] = list2

        return render(request, "app_annotation/consulter_resultats/resultat_corpus.html", {"corpus": corpus, "list": list1, 'categories': categories})
   
    else:
        return HttpResponseRedirect(reverse("login"))     



def annotation_globale(request, corpus_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
       try:
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        categories = corpus.type_annotation.categorie_sentiment_set.all()
        etiq_contrad = corpus.type_annotation.est_contradictoire_set.all()
    
        
        list3 = dict()
        list4 = dict()
        tw_res_type = dict()
        tw_homogene = 0
        tw_heterogene = 0
        tw_contradiction = 0
        tw_non_annote = 0
        for tweet in tweets:
            tous_annnot= Annotation.objects.filter(tweet=tweet)
            list4 = dict()
            for cat in categories:
                votes =0
                for avis in tous_annnot:
                    if avis.avis == cat.id:
                        votes += 1
                list4[cat.id] = votes
            not_zero = 0
            is_zero = 0
            for value in list4.values():
                if(value > 0):
                    not_zero += 1
                if(value==0):
                    is_zero += 1
                
            if (not_zero >= 2):
                est_contr = False
                for contr in etiq_contrad:
                    if (list4[contr.cat1.id] > 0 and list4[contr.cat2.id] > 0 ):
                        tw_res_type[tweet.id] = "Contradiction"
                        est_contr = True
                if (not est_contr):
                    tw_res_type[tweet.id] = "Hétérogène"

                    
                
            else:
                if is_zero == len(list4.values()):
                  tw_res_type[tweet.id] ="Non Annoté"
                else:
                  tw_res_type[tweet.id] ="Homogène"

            list3[tweet.id] = list4
        
        #Calculer % d'Homoginité
        for value in tw_res_type.values():
            if value == "Homogène":
                tw_homogene += 1
        tw_homogene =  "{:.2f}".format((tw_homogene/len(tweets) * 100))
        for value in tw_res_type.values():
            if value == "Hétérogène":
                tw_heterogene += 1
        tw_heterogene =  "{:.2f}".format((tw_heterogene/len(tweets) * 100))
        for value in tw_res_type.values():
            if value == "Contradiction":
                tw_contradiction += 1
        tw_contradiction =  "{:.2f}".format((tw_contradiction/len(tweets) * 100))
        for value in tw_res_type.values():
            if value == "Non Annoté":
                tw_non_annote += 1
        tw_non_annote =  "{:.2f}".format((tw_non_annote/len(tweets) * 100))

        for tw_id in list3:
            tweet_ag = Tweet.objects.get(pk=tw_id)
            max = 0
            etiq = categories[0].nom_categorie
            for key, value in list3[tw_id].items():
                 if value >= max:
                    max = value
                    cat = Categorie_Sentiment.objects.get(pk=(int)(key))
                    etiq = cat.nom_categorie
            if(not corpus.annot):
              
              dic_to_list = list(list3[tw_id].values())  
              occur = 0 
              for ele in dic_to_list:
                if ele == max:
                    occur += 1
              if occur < 2:
                tweet_ag.annotation_general = etiq
                tweet_ag.save()
              else:
          
                tweet_ag.annotation_general = "Indeterminé"
                tweet_ag.save()
   
        corpus.masque = True
        corpus.annot = True
        corpus.save()  
        print(tw_res_type)
        return render(request, "app_annotation/annotation_globale.html", {"corpus": corpus, "list": list3, 'categories': categories, "etiq_contrad": etiq_contrad, "tw_res_type": tw_res_type, "tw_homogene": tw_homogene, "tw_heterogene": tw_heterogene, "tw_contradiction": tw_contradiction, "tw_non_annote":tw_non_annote })
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')


    else:
        return HttpResponseRedirect(reverse("login")) 



def exporter_resultat(request, corpus_id):

    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:

       try:
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        response = HttpResponse(content_type='application/ms-excel')
        response["Content-Disposition"] = 'attachment; filename=' + corpus.titre_corpus + \
             str("-Resultats")+'.xls'
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Sheet1")
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ["Tweet", "Annotation Generale"]

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        font_style = xlwt.XFStyle()
        rows = tweets.values_list("text", "annotation_general")

        for row in rows:
            row_num +=1

            for col_num in range(len(row)):
                ws.write(row_num, col_num, (str)(row[col_num]), font_style)
        wb.save(response)
        return response
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
    else:
        return HttpResponseRedirect(reverse("login"))
    


def corpus_resultat_graph(request, corpus_id):

    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
       try: 
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets= corpus.tweet_set.all()
        categories = corpus.type_annotation.categorie_sentiment_set.all()
        dict_avis = dict()
        label = []
        data_pie = []  
        data_bar = []
        indetermine = 0
        for cat in categories:
            votes = 0
            for tweet in tweets:
                if tweet.annotation_general == (str)(cat.nom_categorie):
                    votes += 1
            label.append(cat.nom_categorie)
            data_bar.append(votes)
            votes =  "{:.2f}".format((votes/ tweets.count()) * 100)
            data_pie.append(votes)
        for tweet in tweets:
            if tweet.annotation_general == "Indeterminé":
                indetermine += 1
        label.append("Indeterminé")
        data_bar.append(indetermine)
        indetermine =  "{:.2f}".format((indetermine /tweets.count()) * 100)
        data_pie.append(indetermine) 

        
        return render(request, "app_annotation/resultat_graph.html", {"label": label, "data_pie": data_pie, "data_bar": data_bar})
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
    else:
        return HttpResponseRedirect(reverse("login"))


def chercher_tweets(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        if request.method == "POST":
         try:
            keyword = request.POST["keyword"]
            limit = request.POST["limit"]
            start_date = request.POST["start_date"]
            end_date = request.POST["end_date"]
            if start_date == "" and end_date == "":
                query = keyword + " lang:ar"
                print (query)
            elif end_date == "":
                query = keyword + " lang:ar "  + "since:" + start_date
            elif start_date == "":
                query = keyword + " lang:ar until:" + end_date
            else:
                query = keyword + " lang:ar until:" + end_date + "since:" + start_date

            return HttpResponseRedirect(reverse("chercher_tweets_keyword", args=(query,limit)))
         except:
             return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
          
        else:
          try:
            tous_corpus = Corpus.objects.all()
            return render(request, "app_annotation/chercher_tweets.html", {"tous_corpus": tous_corpus})
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
          
    else:
        return HttpResponseRedirect(reverse("login"))  

def chercher_tweets_keyword(request, query, limit):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        if request.method == "POST":
          try:
            tw_post = request.POST.getlist("tweet")
            corpus_post = request.POST["corpus"]
            corpus = Corpus.objects.get(pk=corpus_post)
            
            for tweet in tw_post:
                tw = Tweet(text = tweet, corpus = corpus)
                tw.save()
            return HttpResponseRedirect(reverse("modifier_corpus_id",args=(corpus.id,)))
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
           
        else:
          try:
            tous_corpus = Corpus.objects.all()
            query_keyword = query
            tweets = []
        
            for tweet in sntwitter.TwitterSearchScraper(query_keyword).get_items():
               if len(tweets) == limit:
                   break
               else:
                   tweets.append([tweet.date, tweet.content])
   
            
            return render(request, "app_annotation/chercher_tweets_keyword.html", {"tous_corpus": tous_corpus, "tweets": tweets, "query": query, "limit": limit})
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
          
    else:
        return HttpResponseRedirect(reverse("login"))  
 #Gerer compte Admin
    
def gerer_contr(request, contr_id):

    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
          try:  
            compte_cont = Compte.objects.get(pk=contr_id)
            contr = Contributeur.objects.get(pk=compte_cont)
            return render(request, "app_annotation/gerer_compte_contr/gerer_contr.html", {"contr": contr})
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
    else:
        return HttpResponseRedirect(reverse("login"))
    


#Navigation

def corpus(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        tous_corpus = Corpus.objects.all()
        return render(request, "app_annotation/gerer_corpus/corpus.html", {"tous_corpus": tous_corpus})
    else:
        return HttpResponseRedirect(reverse("login"))  


def type_annotation(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        tous_type = Type_Annotation.objects.all()
        return render(request, "app_annotation/gerer_type_annot/type_annotation.html", {"tous_type": tous_type})
    else:
        return HttpResponseRedirect(reverse("login"))  
    
def resultat(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        tous_corpus = Corpus.objects.all()
        list = []
        for corpus in tous_corpus:
            nb_contr = 0
            for cont in Contributeur.objects.all():
                contribution = Annotation.objects.filter(contributeur = cont, corpus = corpus)
                if contribution:
                    nb_contr += 1
            list.append(nb_contr)

        return render(request, "app_annotation/consulter_resultats/resultat.html", {"tous_corpus": tous_corpus, "list": list})
    else:
        return HttpResponseRedirect(reverse("login"))  

def contributeurs(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        tous_contr = Contributeur.objects.all()
        return render(request, "app_annotation/gerer_compte_contr/contributeurs.html", {"tous_contr": tous_contr})
    else:
        return HttpResponseRedirect(reverse("login"))    

def mes_annotations(request):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
        if request.method == "POST":
            pass
        else:
            tous_corpus = Corpus.objects.all()
            tous_annot = Annotation.objects.filter(contributeur=compte_session)
            tous_corpus_annot = []
            liste = []
            for corpus in tous_corpus:
                for annot in tous_annot:
                     if annot.corpus == corpus:
                          tous_corpus_annot.append(corpus)
                          nb_vote = 0
                          annotations = Annotation.objects.filter(corpus=corpus, contributeur = compte_session)
                          for annot in annotations:
                             nb_vote += 1
                          liste.append(nb_vote)  
                          break
                          
            
             
            return render(request, "app_annotation/mes_annotations.html", {"tous_corpus": tous_corpus_annot,"tous_annot": tous_annot, "list": liste})
    else:
        return HttpResponseRedirect(reverse("login"))


def mes_annotations_id(request, corpus_id):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
       try: 
        corpus = Corpus.objects.get(pk=corpus_id)
        categories = corpus.type_annotation.categorie_sentiment_set.all()
        annot_corpus_contr = Annotation.objects.filter(corpus=corpus, contributeur = compte_session)
        liste = dict()
        liste2 = dict()
        for tweet in corpus.tweet_set.all():
            annot = Annotation.objects.filter(tweet=tweet, contributeur = compte_session)
            liste2 = dict()
            for cat in categories:
              if annot:  
                if annot[0].avis == cat.id:
                     liste2[cat.id] = "\u2716"

                else:
                     liste2[cat.id] = "-"
              else:
                for cat in categories:
                    liste2[cat.id] = "-"
                  
            liste[tweet.id] = liste2

        return render(request, "app_annotation/mes_annotations_id.html", {"corpus": corpus, "list": liste})
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
    else:
        return HttpResponseRedirect(reverse("login"))

def mon_compte(request):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
        if request.method == "POST":
           
            prenom = request.POST["prenom"]
            nom = request.POST["nom"]
            tele = request.POST["tele"]
            email = request.POST["email"]
            profession = request.POST["profession"]
            contr = Contributeur.objects.get(pk=compte_session)
            contr.prenom = prenom
            contr.nom = nom
            if tele != "":
               contr.tele = tele
            else:
                contr.tele = None
            contr.email = email
            contr.profession = profession
           
            contr.save()
            return HttpResponseRedirect(reverse("mon_compte_done"))
        
        else:
            
            contr = Contributeur.objects.get(pk=compte_session)
             
            return render(request, "app_annotation/mon_compte.html", {"contr": contr})
    else:
        return HttpResponseRedirect(reverse("login"))
    

def bloquer(request, contr_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
       try: 
        compte_cont = Compte.objects.get(pk=contr_id)
        contr = Contributeur.objects.get(pk=compte_cont)
        if contr.bloque:
            contr.bloque = False
            contr.save()
        else:
            contr.bloque = True
            contr.save()
        return render(request, "app_annotation/gerer_compte_contr/gerer_contr.html", {"contr": contr})
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
    else:
        return HttpResponseRedirect(reverse("login"))  




def upload_corpus_done(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        corpus_titre = request.POST["corpus_titre"]
        corpus_file = request.FILES['file']
        masque = request.POST["masque"]
        type_annot_post = request.POST["type_annot"]
        type_annot = Type_Annotation.objects.get(pk=(int)(type_annot_post))
        try: 
          workbook = load_workbook(filename=corpus_file.file)
          Page = workbook['Sheet1']
          column = Page["A"]
        except:
          form_redir = UploadFileForm()
          types_annot_redir = Type_Annotation.objects.all()
          return HttpResponseRedirect(reverse("upload_corpus_error"))
        corpus = Corpus(titre_corpus=corpus_titre, type_annotation=type_annot, masque=(int)(masque))
        corpus.save()
        for cell in column:
            if cell.value!=None:
              cell_txt =cell.value
              tweet = Tweet(text=cell_txt, corpus=corpus)
              tweet.save()
            else:
                continue
        form_redir = UploadFileForm()
        types_annot_redir = Type_Annotation.objects.all()
        return HttpResponseRedirect(reverse("upload_corpus_done"))
      else:
        form = UploadFileForm()
        types_annot = Type_Annotation.objects.all()
        return render(request, "app_annotation/gerer_corpus/upload.html", {"form": form,  "types_annot": types_annot, "cree": True})
    else:
        return HttpResponseRedirect(reverse("login"))  
    

def upload_corpus_error(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        corpus_titre = request.POST["corpus_titre"]
        corpus_file = request.FILES['file']
        masque = request.POST["masque"]
        type_annot_post = request.POST["type_annot"]
        type_annot = Type_Annotation.objects.get(pk=(int)(type_annot_post))
        try: 
          workbook = load_workbook(filename=corpus_file.file)
          Page = workbook['Sheet1']
          column = Page["A"]
        except:
          form_redir = UploadFileForm()
          types_annot_redir = Type_Annotation.objects.all()
          return HttpResponseRedirect(reverse("upload_corpus_error"))
        corpus = Corpus(titre_corpus=corpus_titre, type_annotation=type_annot, masque=(int)(masque))
        corpus.save()
        for cell in column:
            if cell.value!=None:
              cell_txt =cell.value
              tweet = Tweet(text=cell_txt, corpus=corpus)
              tweet.save()
            else:
                continue
        form_redir = UploadFileForm()
        types_annot_redir = Type_Annotation.objects.all()
        return HttpResponseRedirect(reverse("upload_corpus_done"))
      else:
        form = UploadFileForm()
        types_annot = Type_Annotation.objects.all()
        return render(request, "app_annotation/gerer_corpus/upload.html", {"form": form,  "types_annot": types_annot, "error": True})
    else:
        return HttpResponseRedirect(reverse("login"))  


def modifier_corpus_id_done(request, corpus_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
           if request.method == "POST":
              try:
                corpus = Corpus.objects.get(pk=corpus_id)
                corpus_titre = request.POST["corpus_titre"]
                type_annot_post = request.POST["type_annot"]
                type_annot = Type_Annotation.objects.get(pk=type_annot_post)
                type_annot_old = corpus.type_annotation
                tweets_post = request.POST.getlist("tweet")
                corpus_old_tweets = corpus.tweet_set.all()
             
            
                if (len(tweets_post) == len(corpus_old_tweets)) and ((int)(type_annot_post) == type_annot_old.id):
                    same = 0
                    for i, tweet in enumerate(tweets_post):
                        if tweet == (str)(corpus_old_tweets[i].text):
                            same += 1
                    if same == len(corpus_old_tweets):
                            pass
                    else:
                        for tweet in corpus_old_tweets:
                            tweet.delete()

                        for tweet in tweets_post:
                            tweet_db = Tweet(text = tweet, corpus=corpus)
                            tweet_db.save()
                            
                else:
                    for tweet in corpus_old_tweets:
                        tweet.delete()

                    for tweet in tweets_post:
                        tweet_db = Tweet(text = tweet, corpus=corpus)
                        tweet_db.save()
            
                corpus.type_annotation = type_annot
                corpus.titre_corpus = corpus_titre
                corpus.save()   
                return HttpResponseRedirect(reverse("modifier_corpus_id_done"), args=(corpus_id,))
              except:
                return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
           else:
              try: 
               corpus = Corpus.objects.get(pk=corpus_id)
               types_annot = Type_Annotation.objects.all()
               return render(request, "app_annotation/gerer_corpus/modifier_corpus_id.html", {"corpus": corpus, "types_annot": types_annot, "cree": True})
              except:
                  return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
           

def ajouter_type_annotation_done(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        type_annotation_post = request.POST["type_annotation"]
        categories = request.POST.getlist("categorie")
        type_annotation = Type_Annotation(nom_type=type_annotation_post)
        type_annotation.save()
        for cat in categories:
            cat_db = Categorie_Sentiment(nom_categorie=cat, type_annotation=type_annotation)
            cat_db.save()
        return HttpResponseRedirect(reverse("ajouter_type_annotation_done")) 
      else:
        return render(request, "app_annotation/gerer_type_annot/ajouter_type_annotation.html", {"cree": True})
    else:
        return HttpResponseRedirect(reverse("login")) 


def ajouter_etiquette_done(request):
        
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
            if request.method == "POST":
                nom = request.POST["etiquette"]
                type_post = request.POST["type_annot"]
                type_annot = Type_Annotation.objects.get(pk=type_post)
                etiquette = Categorie_Sentiment(nom_categorie=nom, type_annotation = type_annot)
                etiquette.save()
                types_annot_red = Type_Annotation.objects.all()
                return HttpResponseRedirect(reverse("ajouter_etiquette_done"))
            else:
        
               types_annot = Type_Annotation.objects.all()
               return render(request, "app_annotation/gerer_type_annot/ajouter_etiquette.html", {"types_annot": types_annot, "cree": True})
        
    else:
        return HttpResponseRedirect(reverse("login")) 


def modifier_type_annotation_id_done(request, type_id):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
        if request.method == "POST":
           try: 
            type_annotation_post = request.POST["type_annotation"]
            categories = request.POST.getlist("categorie")
            type_annot = Type_Annotation.objects.get(pk=type_id)
            for cat in type_annot.categorie_sentiment_set.all():
                cat.delete()
            for cat in categories:
                cat_db = Categorie_Sentiment(nom_categorie=cat, type_annotation=type_annot)
                cat_db.save()
            type_annot.nom_type = type_annotation_post
            type_annot.save()

            type_annot_red = Type_Annotation.objects.get(pk=type_id)
            categories_red = type_annot.categorie_sentiment_set.all()
            return HttpResponseRedirect(reverse("modifier_type_annotation_id_done", args=(type_id, )))
           except:
               return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')

        else:
           try: 
            type_annot = Type_Annotation.objects.get(pk=type_id)
            categories = type_annot.categorie_sentiment_set.all()
            return render(request, "app_annotation/gerer_type_annot/modifier_type_annotation_id.html", {"type_annot": type_annot, "categories": categories, "cree": True})
           except:
               return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')


def creer_corpus_done(request):
    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
        corpus_titre = request.POST["corpus_titre"]
        masque = request.POST["masque"]
        type_annot_post = request.POST["type_annot"]
        type_annot = Type_Annotation.objects.get(pk=(int)(type_annot_post))
        tweets_post = request.POST.getlist("tweet")
        corpus = Corpus(titre_corpus=corpus_titre, type_annotation=type_annot, masque=(int)(masque))
        corpus.save()

        for tweet in tweets_post:
            tweet_db = Tweet(text=tweet, corpus=corpus)
            tweet_db.save()
        types_annot_redir = Type_Annotation.objects.all()
        cree = True
        return HttpResponseRedirect(reverse("creer_corpus_done"))
      else:
        types_annot = Type_Annotation.objects.all()
        return render(request, "app_annotation/gerer_corpus/creer_corpus.html", {"types_annot": types_annot, "cree": True})
    else:
        return HttpResponseRedirect(reverse("login")) 
    



def annoter_corpus_done(request, corpus_id):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
       if request.method == "POST":
          try: 
           corpus = Corpus.objects.get(pk=corpus_id)
           tweets = corpus.tweet_set.all()
           tous_avis = request.POST.getlist("avis") 
           
           tous_avis_old= Annotation.objects.filter(corpus= corpus, contributeur = compte_session)
           for avis in tous_avis_old:
               avis.delete()
           
           for i, avis in enumerate(tous_avis):
              if avis != "":
                 
                   avis_split = avis.split("_")
                   cat_id = avis_split[0]
                   tw_id = avis_split[1]
                   tw = Tweet.objects.get(pk=(int)(tw_id))
                   annot = Annotation(avis=cat_id, tweet = tw, contributeur= compte_session, corpus=corpus)
                   annot.save()
           


           corpus_red = Corpus.objects.get(pk=corpus_id)
           tweets_red = corpus_red.tweet_set.all()
           tous_avis_red = Annotation.objects.filter(corpus=corpus_red, contributeur=compte_session)
           type_annot_red = corpus_red.type_annotation
           categories_red = type_annot_red.categorie_sentiment_set.all()
           return HttpResponseRedirect(reverse("annoter_corpus_done", args=(corpus_id, )))
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
       else:
          try:
           corpus = Corpus.objects.get(pk=corpus_id)
           tweets = corpus.tweet_set.all()
           tous_avis = Annotation.objects.filter(corpus=corpus, contributeur=compte_session)
           type_annot = corpus.type_annotation
           categories = type_annot.categorie_sentiment_set.all()
           return render(request, "app_annotation/consulter_corpus/annoter_corpus.html", {"corpus": corpus, "categories": categories, "tweets": tweets, "tous_avis": tous_avis, "contributeur_login": compte_session.login.login, "cree" : True })
          except:
              return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
       

def annotation_globale_modifier(request, corpus_id):

    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
       try: 
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        tous_ag = request.POST.getlist("ag_tweet")

        for tw, ag in zip(tweets, tous_ag):
            print(ag)
            tw.annotation_general = ag
            tw.save()
           
        

        return HttpResponseRedirect(reverse("annotation_globale_modifier_done", args=(corpus_id, )))
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
         
      else:
       try: 
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        categories = corpus.type_annotation.categorie_sentiment_set.all()
        etiq_contrad = corpus.type_annotation.est_contradictoire_set.all()

        tw_res_type = dict()
        tw_homogene = 0
        tw_heterogene = 0
        tw_contradiction = 0
        tw_non_annote = 0


        list3 = dict()
        list4 = dict()

   
        for tweet in tweets:
            tous_annnot= Annotation.objects.filter(tweet=tweet)
            list4 = dict()
            for cat in categories:
                votes =0
                for avis in tous_annnot:
                    if avis.avis == cat.id:
                        votes += 1
                list4[cat.id] = votes
            list3[tweet.id] = list4
            not_zero = 0
            is_zero = 0
            for value in list4.values():
                if(value > 0):
                    not_zero += 1
                if(value==0):
                    is_zero += 1
                
            if (not_zero >= 2):
                est_contr = False
                for contr in etiq_contrad:
                    if (list4[contr.cat1.id] > 0 and list4[contr.cat2.id] > 0 ):
                        tw_res_type[tweet.id] = "Contradiction"
                        est_contr = True
                if (not est_contr):
                    tw_res_type[tweet.id] = "Hétérogène"
            else:
                if is_zero == len(list4.values()):
                  tw_res_type[tweet.id] ="Non Annoté"
                else:
                  tw_res_type[tweet.id] ="Homogène"


        #Calculer % d'Homoginité
        for value in tw_res_type.values():
            if value == "Homogène":
                tw_homogene += 1
        tw_homogene =  "{:.2f}".format((tw_homogene/len(tweets) * 100))
        for value in tw_res_type.values():
            if value == "Hétérogène":
                tw_heterogene += 1
        tw_heterogene =  "{:.2f}".format((tw_heterogene/len(tweets) * 100))
        for value in tw_res_type.values():
            if value == "Contradiction":
                tw_contradiction += 1
        tw_contradiction =  "{:.2f}".format((tw_contradiction/len(tweets) * 100))
        for value in tw_res_type.values():
            if value == "Non Annoté":
                tw_non_annote += 1
        tw_non_annote =  "{:.2f}".format((tw_non_annote/len(tweets) * 100))
   


        return render(request, "app_annotation/annotation_globale_modifier.html", {"corpus": corpus, "list": list3, 'categories': categories, "etiq_contrad": etiq_contrad, "tw_res_type": tw_res_type, "tw_homogene": tw_homogene, "tw_heterogene": tw_heterogene, "tw_contradiction": tw_contradiction, "tw_non_annote":tw_non_annote})
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')


    else:
        return HttpResponseRedirect(reverse("login")) 
    



def annotation_globale_modifier_done(request, corpus_id):

    try:
         compte_session = Compte.objects.get(pk=request.session["login_admin"])
    except (KeyError, Compte.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.password == request.session["password_admin"]:
      if request.method == "POST":
       try: 
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        tous_ag = request.POST.getlist("ag_tweet")

        for tw, ag in zip(tweets, tous_ag):
            print(ag)
            tw.annotation_general = ag
            tw.save()
           
        

        return HttpResponseRedirect(reverse("annotation_globale_modifier_done", args=(corpus_id, )))
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')
         
      else:
       try: 
        corpus = Corpus.objects.get(pk=corpus_id)
        tweets = corpus.tweet_set.all()
        categories = corpus.type_annotation.categorie_sentiment_set.all()
        etiq_contrad = corpus.type_annotation.est_contradictoire_set.all()

        tw_res_type = dict()
        tw_homogene = 0
        tw_heterogene = 0
        tw_contradiction = 0
        tw_non_annote = 0


        list3 = dict()
        list4 = dict()

   
        for tweet in tweets:
            tous_annnot= Annotation.objects.filter(tweet=tweet)
            list4 = dict()
            for cat in categories:
                votes =0
                for avis in tous_annnot:
                    if avis.avis == cat.id:
                        votes += 1
                list4[cat.id] = votes
            list3[tweet.id] = list4
            not_zero = 0
            is_zero = 0
            for value in list4.values():
                if(value > 0):
                    not_zero += 1
                if(value==0):
                    is_zero += 1
                
            if (not_zero >= 2):
                est_contr = False
                for contr in etiq_contrad:
                    if (list4[contr.cat1.id] > 0 and list4[contr.cat2.id] > 0 ):
                        tw_res_type[tweet.id] = "Contradiction"
                        est_contr = True
                if (not est_contr):
                    tw_res_type[tweet.id] = "Hétérogène"
            else:
                if is_zero == len(list4.values()):
                  tw_res_type[tweet.id] ="Non Annoté"
                else:
                  tw_res_type[tweet.id] ="Homogène"


        #Calculer % d'Homoginité
        for value in tw_res_type.values():
            if value == "Homogène":
                tw_homogene += 1
        tw_homogene = (tw_homogene/len(tweets) * 100)
        for value in tw_res_type.values():
            if value == "Hétérogène":
                tw_heterogene += 1
        tw_heterogene = (tw_heterogene/len(tweets) * 100)
        for value in tw_res_type.values():
            if value == "Contradiction":
                tw_contradiction += 1
        tw_contradiction = (tw_contradiction/len(tweets) * 100)
        for value in tw_res_type.values():
            if value == "Non Annoté":
                tw_non_annote += 1
        tw_non_annote = (tw_non_annote/len(tweets) * 100)
   


        return render(request, "app_annotation/annotation_globale_modifier.html", {"corpus": corpus, "list": list3, 'categories': categories, "etiq_contrad": etiq_contrad, "tw_res_type": tw_res_type, "tw_homogene": tw_homogene, "tw_heterogene": tw_heterogene, "tw_contradiction": tw_contradiction, "tw_non_annote":tw_non_annote, "cree": True})
       except:
           return HttpResponse('<h1 style="margin: 0 auto; width: fit-content">Erreur Survenue<h1>')


    else:
        return HttpResponseRedirect(reverse("login")) 
    

def mon_compte_done(request):
    try:
         compte_session = Contributeur.objects.get(pk=request.session["login_contri"])
    except (KeyError, Contributeur.DoesNotExist):
        return HttpResponseRedirect(reverse("login"))
    if compte_session.login.password == request.session["password_contri"]:
        if request.method == "POST":
            login = request.POST["login"]
            prenom = request.POST["prenom"]
            nom = request.POST["nom"]
            email = request.POST["email"]
            profession = request.POST["profession"]
            contr = Contributeur.objects.get(pk=compte_session)
            contr.login = login
            contr.prenom = prenom
            contr.nom = nom
            contr.email = email
            contr.profession = profession
            contr.save()      
            return HttpResponseRedirect(reverse("mon_compte_done"))
        

        else:
            
            contr = Contributeur.objects.get(pk=compte_session)
             
            return render(request, "app_annotation/mon_compte.html", {"contr": contr, "cree": True})
    else:
        return HttpResponseRedirect(reverse("login"))